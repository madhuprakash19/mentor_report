from django.shortcuts import render
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from marks.models import (mentor_list,subject,
                mentor_data,student, user_profile)

from django.http import HttpResponseRedirect
from django.urls import reverse

import openpyxl

def faculty_status(user):
    teacher = user_profile.objects.get(user=user)
    if teacher.role_id.id == 2:
        return True
    else:
        return False

def mentor(request):
    teacher = request.user
    student = list(mentor_list.objects.filter(faculty=teacher))
    return render(request,'mentor.html',{'student':student})

def marksedit(request,id):
    temp = mentor_data.objects.get(id=id)
    teacher = request.user

    if temp.student_meta.faculty != teacher:
        return render(request,'403.html',)

    if request.method == 'POST':
        mse1 = request.POST['mse1_marks']
        mse2 = request.POST['mse2_marks']
        mse3 = request.POST['mse3_marks']
        la1 = request.POST['la1']
        la2 = request.POST['la2']
        final = request.POST['final_cie']
        id1 = request.POST['mentor_data_id']

        emarks = mentor_data.objects.get(id=id1)
        emarks.mse1_marks = mse1
        emarks.mse2_marks = mse2
        emarks.mse1_marks = mse1
        emarks.la1 = la1
        emarks.la1 = la1
        emarks.final_cie = final
        emarks.save()
        id = emarks.student_meta.id

        return HttpResponseRedirect(reverse('marks:report',args=[id]))


    return render(request,'marksedit.html',{'temp':temp})

def iamarks(request):
    if "GET" == request.method:
        return render(request, 'iamarks.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        print(type(excel_file))
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Book1"]
        excel_data = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        #check for total columns
        if len(excel_data[0]) != 3:
            return render(request, 'iamarks.html', {"sucess":"File not in proper format"})

        for i in range (1,len(excel_data)):
            #check for values
            usn = excel_data[i][0]
            marks = excel_data[i][1]
            sub_ = excel_data[i][2]

            if len(usn)!=10:
                return render(request, 'iamarks.html', {"sucess":"File not in proper format"})

        for i in range (1,len(excel_data)):
            usn = excel_data[i][0]
            marks = excel_data[i][1]
            sub_ = excel_data[i][2]

            try:
                student_ = User.objects.get(username=usn)
                try:
                    temp =  mentor_list.objects.get(student=student_)
                    try:
                        sub = subject.objects.get(code=sub_)
                        md = mentor_data.objects.get(student_meta=temp,subject=sub)
                        md.mse1_marks = marks
                        md.save()
                    except:
                        sub = subject.objects.get(code=sub_)
                        md = mentor_data(student_meta=temp,subject=sub,mse1_marks=marks)
                        md.save()
                except:
                    print("student in mentor not found")
            except:
                print("student not found")

        return render(request, 'iamarks.html', {"excel_data":excel_data,"sucess":"data saved"})

@login_required()
def home(request):
    temp = list(student.objects.all())
    print(temp)
    return render(request,'index.html',{'temp':temp})

def mapping(request):
    if "GET" == request.method:
        return render(request, 'mapping.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Book1"]
        print(worksheet)
        excel_data = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        if len(excel_data[0]) != 5:
            return render(request, 'mapping.html', {"sucess":"File not in proper format"})
        for i in range (1,len(excel_data)):
            usn = excel_data[i][0]
            mentor = excel_data[i][1]
            sem_ = excel_data[i][2]
            section_ = excel_data[i][3]
            ay_ = excel_data[i][4]

            if len(usn)!=10 or len(str(sem_))!=1 or len(section_)!=1 or len(ay_)!=9:
                return render(request, 'mapping.html', {"sucess":"File not in proper format"})



        for i in range (1,len(excel_data)):
            usn = excel_data[i][0]
            mentor = excel_data[i][1]
            sem_ = excel_data[i][2]
            section_ = excel_data[i][3]
            ay_ = excel_data[i][4]

            try:
                student=User.objects.create_user(usn, password=usn)
                t = User.objects.create_user(mentor, password=mentor)
                student.save()
                print(usn+" created")
            except:
                print(usn+" already Exists")
                pass

            student_ = User.objects.get(username=usn)
            teacher_ = User.objects.get(username=mentor)

            try:
                temp =  mentor_list.objects.get(student=student_,ay=ay_)
                temp.faculty = teacher_
                temp.save()
            except:
                temp = mentor_list(student=student_,faculty=teacher_,sem=sem_,section=section_,ay=ay_)
                temp.save()



        return render(request, 'mapping.html', {"excel_data":excel_data,"sucess":"data saved"})

def report(request,id):
    stu = mentor_list.objects.get(id=id)
    teacher = request.user
    flag = 1
    temp = list(mentor_data.objects.filter(student_meta=stu))
    for i in temp:
        if i.student_meta.faculty != teacher:
            flag = 0
    return render(request,'report.html',{'temp':temp,'flag':flag})
